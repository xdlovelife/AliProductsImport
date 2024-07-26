import logging
import time
import tkinter as tk
from tkinter import filedialog
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, NoSuchWindowException
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)

chrome_driver_path = 'D:\\chromedriver-win64\\chromedriver.exe'
chromium_binary_path = 'C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe'

# 用户数据目录路径（保存已安装插件和用户配置）
user_data_dir = 'C:\\Users\\Administrator\\AppData\\Local\\Google\\Chrome\\User Data'  # 请根据实际路径进行调整

options = Options()
options.binary_location = chromium_binary_path
options = webdriver.ChromeOptions()


# 指定用户数据目录
options.add_argument(f'--user-data-dir={user_data_dir}')


def open_browser():
    attempts = 3
    for attempt in range(1, attempts + 1):
        try:
            driver = webdriver.Chrome(service=Service(chrome_driver_path), options=options)
            logger.info("Chrome WebDriver启动成功。")
            return driver
        except Exception as e:
            logger.error(f"第 {attempt} 次尝试启动Chrome失败: {str(e)}")
            if attempt == attempts:
                logger.error("达到最大重试次数。退出程序。")
                return None
            time.sleep(3)  # 等待后重试
    return None


def open_alibaba(driver, selected_categories, sheet_name):
    try:
        if driver:
            url = "https://www.alibaba.com/"
            logger.info(f"访问页面: {url}")
            driver.get(url)
            # 等待 fy23-icbu-search-bar-inner 元素加载完成
            search_bar = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'fy23-icbu-search-bar-inner'))
            )

            for category in selected_categories:
                try:
                    process_link(driver, "https://www.alibaba.com/", category, sheet_name)
                except Exception as e:
                    logger.error(f"处理类别 '{category}' 出错: {e}")
                    # Handle specific errors or retry logic here if needed

            driver.quit()

    except NoSuchElementException as e:
        logger.error(f"未找到元素：{e}")
    except TimeoutException as e:
        logger.error(f"超时等待元素加载：{e}")
    except Exception as e:
        logger.error(f"发生异常: {str(e)}")


def process_link(driver, link, category, sheet_name):
    try:
        logger.info(f"处理分类: {category}")
        logger.info(f"处理链接: {link}")
        driver.get(link)
        driver.switch_to.window(driver.window_handles[0])

        # 等待搜索框加载完成
        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input.search-bar-input.util-ellipsis'))
        )
        # 将产品分类名称填入搜索框
        search_input.clear()
        search_input.send_keys(category)

        # 点击搜索按钮
        search_button = driver.find_element(By.CSS_SELECTOR, 'button.fy23-icbu-search-bar-inner-button')
        search_button.click()

        # 等待产品列表加载完成
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.CLASS_NAME, "organic-list"))
        )

        # 模拟向下滚动页面，直到加载完所有产品
        last_height = driver.execute_script("return document.body.scrollHeight")
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)  # 等待加载
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        # 加载完所有产品后，获取产品列表的长度并打印
        product_list = driver.find_elements(By.CLASS_NAME, "fy23-search-card")
        num_products = len(product_list)
        logger.info(f"共抓取到的产品数量：{num_products}")

        # 循环处理产品
        success_count = 0
        for product in product_list:
            start_time = time.time()  # 记录处理开始时间

            try:
                # 获取产品标题
                product_title = product.find_element(By.CLASS_NAME, "search-card-e-title")
                logger.info(f"当前产品标题: {product_title.text}")

                # 滚动到产品标题所在位置
                scroll_to_element(driver, product_title)
                time.sleep(1)

                # 获取产品链接并打开
                product_link = product.find_element(By.TAG_NAME, "a").get_attribute("href")
                driver.execute_script(f"window.open('{product_link}')")

                # 处理产品详情页操作
                success_count = handle_product_detail(driver, category, success_count, sheet_name)

                # 等待一段时间，可以根据实际情况调整
                time.sleep(1)

                end_time = time.time()  # 记录处理结束时间
                processing_time = end_time - start_time  # 计算处理时间
                logger.info(f"处理时间：{processing_time:.2f} 秒")  # 输出处理时间

                # 获取屏幕宽度（如果无法获取，则使用默认值）
                screen_width = get_screen_width()
                # 打印分割线
                logger.info("=" * (screen_width // 10))  # 根据屏幕宽度计算分割线长度

                success_count += 1
            except NoSuchElementException as e:
                logger.error(f"未找到产品标题或链接: {e}")
            except Exception as e:
                logger.error(f"处理产品时发生错误: {e}")

        logger.info(f"成功处理的产品数量: {success_count}")
        return success_count

    except Exception as e:
        logger.error(f"处理链接时发生错误: {e}")
        return 0


def handle_product_detail(driver, category, success_count, sheet_name):
    try:
        # 获取所有窗口句柄
        original_window = driver.current_window_handle
        handles = driver.window_handles

        # 获取新打开的窗口句柄
        new_window = None
        for window_handle in handles:
            if window_handle != original_window:
                new_window = window_handle
                break

        if new_window:
            # 切换到新打开的产品详情页窗口
            driver.switch_to.window(new_window)
            # 等待产品详情页元素加载完成
            product_title = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.TAG_NAME, "h1")))
            # 处理产品详情页操作，这里可以根据实际需要修改
            success_count = handle_product_actions(driver, category, success_count, sheet_name)
            # 等待一段时间，可以根据实际情况调整
            time.sleep(1)
            # 切换回原始窗口（产品搜索页）
            driver.switch_to.window(original_window)
        return success_count
    except NoSuchWindowException as e:
        logger.error(f"浏览器窗口丢失：{e}")
        return success_count
    except Exception as e:
        logger.error(f"处理产品详情页时发生错误: {e}")
        close_tab(driver, new_window)  # 关闭出错的产品详情页标签页
        return success_count


from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time


def fetch_dropdown_options(driver, sheet_name):
    try:
        # 找到下拉菜单的整个区域
        dropdown = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.CLASS_NAME, 'ms-drop')))

        # 找到搜索框并输入关键词（如果有的话）
        search_box = dropdown.find_element(By.CSS_SELECTOR, '.ms-search input[type="text"]')
        search_box.clear()
        search_box.send_keys(sheet_name)  # 替换为你想搜索的关键词

        # 等待搜索结果更新，这里可以根据具体情况增加等待时间
        WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.CLASS_NAME, 'ms-ajax-search-loader')))

        # 找到所有的li元素（选项）
        options = dropdown.find_elements(By.XPATH, '//ul/li[not(@class="ms-no-results") and not(@class="ms-ajax-search-loader")]')

        for option in options:
            # 获取每个选项的文本内容
            option_text = option.text.strip()
            print(f"当前 li 内容: {option_text}")

    except Exception as e:
        print(f"发生异常: {str(e)}")



def handle_product_actions(browser, category, success_count, sheet_name):
    logger.info(f"处理产品详情页操作: {category}, {sheet_name}!!!")
    try:
        add_btn_con = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="addBtnCon"]')))
        add_btn_con.click()
        logging.info("点击了按钮//*[@id='addBtnCon']")

        try:
            element = WebDriverWait(browser, 20).until(
                EC.presence_of_element_located((By.XPATH, '//span[@class="inactive" and text()="Draft"]'))
            )
            logging.info("成功加载 Draft 元素")
            actions = ActionChains(browser)
            actions.move_to_element(element).perform()
            element.click()
            logging.info("成功点击 Draft 元素")
            time.sleep(2)
        except Exception as e:
            logging.error(f"等待和点击 Draft 元素时出现错误：{e}")
            close_current_tab(browser)
            return success_count

        time.sleep(3)  # 可以根据实际情况调整等待时间

        # 检查是否出现 "This product is already in your store, what would you like to do?"
        success_message = None
        try:
            success_message = browser.find_element(By.XPATH,
                                                   '//div[@class="textcontainer centeralign home-content "]/p[1]')
            if success_message.text == "This product is already in your store, what would you like to do?":
                logging.info("产品已存在，不再处理当前产品")
            browser.close()  # 关闭当前产品详情页标签页
            return success_count  # 跳出函数，不再处理当前产品
        except NoSuchElementException:
            pass  # 如果未找到消息元素，继续后续操作

        time.sleep(2)  # 可以根据实际情况调整等待时间

        # 继续后续操作，例如选择下拉菜单中的类别等
        select_button = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@class="ms-choice"]'))
        )
        logging.info("等待并点击选择按钮")
        select_button.click()

        dropdown = WebDriverWait(browser, 10).until(
            EC.visibility_of_element_located((By.CLASS_NAME, "ms-drop"))
        )

        # # 从 Excel 文件中读取工作表名称作为搜索关键词
        search_input = dropdown.find_element(By.CSS_SELECTOR, ".ms-search input[type='text']")
        search_input.clear()
        search_input.send_keys(sheet_name)
        logging.info(f"输入关键词: {sheet_name}")

        WebDriverWait(browser, 10)

        # 选择匹配的复选框
        fetch_dropdown_options(browser, sheet_name)
        time.sleep(3)

        try:
            # 点击 description_tab_button 按钮
            description_tab_button = browser.find_element(By.XPATH, '//*[@id="description_tab_button"]')
            description_tab_button.click()
            logging.info("点击了 description_tab_button 按钮")
            time.sleep(3)  # 等待页面加载

            # 点击 Variants 按钮
            variants_button = browser.find_element(By.CSS_SELECTOR,
                                                   'button.accordion-tab[data-actab-group="0"][data-actab-id="2"]')
            variants_button.click()
            logging.info("点击了 Variants 按钮")

            # 选择 Import all variants automatically 单选框
            all_variants_radio = browser.find_element(By.ID, 'all_variants')
            all_variants_radio.click()
            logging.info("选择 Import all variants automatically 单选框")

            time.sleep(3)  # 等待页面反应

            # 选择 Select which variants to include 单选框
            price_switch_radio = browser.find_element(By.ID, 'price_switch')
            price_switch_radio.click()
            logging.info("选择 Select which variants to include 单选框")

            time.sleep(3)  # 等待页面反应
        except Exception as e:
            logging.error(f"点击 Variants 按钮时出现错误：{e}")
            close_current_tab(browser)
            return success_count

        # 点击 Images 按钮
        images_button = browser.find_element(By.XPATH,
                                             '//button[@class="accordion-tab accordion-custom-tab" and @data-actab-group="0" and @data-actab-id="3"]')
        images_button.click()
        logging.info("点击了 Images 按钮")
        time.sleep(3)  # 等待页面反应

        add_to_store_button = browser.find_element(By.ID, 'addBtnSec')
        scroll_to_element(browser, add_to_store_button)

        add_to_store_button.click()
        logging.info("成功点击 Add to your Store 按钮")

        logging.info("等待页面加载完成")

        # 等待导入过程完成，确保 importify-app-container 元素出现
        try:
            wait_for_element_to_appear(browser, By.ID, 'importify-app-container')
            logging.info("产品正在导入中...")

            # 等待成功消息出现
            success_message = None
            timeout = 60  # 设定超时时间
            start_time = time.time()
            while time.time() - start_time < timeout:
                try:
                    success_message = browser.find_element(By.XPATH,
                                                           '//div[@class="textcontainer centeralign home-content "]/p[1]')
                    if success_message.text == "We have successfully created the product page.":
                        logging.info(f"产品导入成功, 共计: {success_count + 1}")
                        success_count += 1
                        break
                    else:
                        logging.warning("产品正在导入中...")
                except Exception as e:
                    logging.warning("未检测到产品成功导入，继续等待...")
                time.sleep(5)  # 每秒检查一次

            if not success_message or success_message.text != "We have successfully created the product page.":
                logging.error("超时：未找到成功创建产品页面的消息")

        except Exception as e:
            logging.error(f"页面加载出错: {e}")
            close_current_tab(browser)

        time.sleep(2)
        close_current_tab(browser)
        return success_count

    except NoSuchWindowException as e:
        logging.error(f"浏览器窗口丢失：{e}")
        close_current_tab(browser)
        return success_count
    except Exception as e:
        logging.error(f"处理产品详情页操作时发生错误: {e}")
        close_current_tab(browser)
        return success_count

def close_current_tab(browser):
    try:
        current_window_handle = browser.current_window_handle
        browser.switch_to.window(current_window_handle)
        browser.close()
        logging.info("已关闭处理出错的标签页。")
    except NoSuchWindowException as e:
        logging.error(f"浏览器窗口丢失：{e}")
    except Exception as e:
        logging.error(f"关闭标签页时发生错误: {e}")
def wait_for_element_to_appear(driver, by, selector, timeout=10):
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, selector))
        )
    except TimeoutException:
        logging.error(f"元素未能在 {timeout} 秒内出现: {selector}")
        raise
def close_tab(driver, window_handle):
    try:
        if window_handle:
            driver.switch_to.window(window_handle)
            driver.close()
            logger.info("关闭出错的产品详情页标签页")
    except Exception as e:
        logger.error(f"关闭标签页时发生错误: {e}")
def get_screen_width():
    try:
        root = tk.Tk()
        screen_width = root.winfo_screenwidth()
        root.destroy()
        return screen_width
    except Exception as e:
        logger.error(f"获取屏幕宽度时出错: {e}")
        return 1000  # 返回默认屏幕宽度


def browse_excel_file():
    root = tk.Tk()
    root.withdraw()  # 隐藏Tk窗口
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    return file_path


def read_categories_from_excel(file_path):
    try:
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        categories = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            category = row[0]
            if category:
                categories.append(category)
        return categories
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return []

def read_sheet_names_from_excel(file_path):
    sheet_name = []
    try:
        wb = load_workbook(filename=file_path)
        sheet_name = wb.sheetnames
    except Exception as e:
        logger.error(f"读取Excel文件时发生错误: {e}")
    return sheet_name

def scroll_to_element(browser, element):
    try:
        WebDriverWait(browser, 60).until(EC.visibility_of(element))
        browser.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        logging.info(f"滚动到元素: {element.text}")
    except Exception as e:
        logging.error(f"滚动到元素时出错: {e}")


def main():
    driver = None  # 初始化 driver 变量，以便在后续代码中访问

    try:
        file_path = browse_excel_file()
        if not file_path:
            logger.error("未选择Excel文件。")
            return

        selected_categories = read_categories_from_excel(file_path)
        if not selected_categories:
            logger.error("未从Excel文件中读取到任何类别。")
            return
        logger.info(f"从Excel文件中读取的要导入的产品名称: {selected_categories}")
        driver = open_browser()
        if not driver:
            logger.error("无法启动浏览器。")
            return

        # 获取工作表名称列表
        sheet_name = read_sheet_names_from_excel(file_path)
        if not sheet_name:
            logger.error("未从Excel文件中读取到任何工作表名称。")
            return

        logger.info(f"从Excel文件中读取的工作表名称: {sheet_name}")

        # 调用 open_alibaba() 函数，并传递 driver、selected_categories 和 sheet_names
        open_alibaba(driver, selected_categories, sheet_name)

    except Exception as e:
        logger.error(f"发生异常: {str(e)}")
        pass


if __name__ == "__main__":
    main()
