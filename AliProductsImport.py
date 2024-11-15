import logging
import time
import tkinter as tk
from tkinter import filedialog
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, NoSuchWindowException
from openpyxl import load_workbook
from contextlib import contextmanager
import contextlib

# 设置日志记录器的配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s || %(message)s')
logger = logging.getLogger(__name__)

# 添加配置常量
CONFIG = {
    'CHROME_DRIVER_PATH': 'D:\\chromedriver-win64\\chromedriver.exe',
    'CHROME_BINARY_PATH': 'C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe',
    'USER_DATA_DIR': 'C:\\Users\\Administrator\\AppData\\Local\\Google\\Chrome\\User Data',
    'WAIT_TIMEOUT': 10,
    'WAIT_TIMEOUT_LONG': 20,
    'SCROLL_WAIT': 1,
    'ANIMATION_WAIT': 0.5,
    'IMPORT_TIMEOUT': 60,
    'RETRY_INTERVAL': 2,
    'MAX_RETRIES': 3
}


# 优化浏览器选项设置
def get_chrome_options():
    options = Options()
    options.binary_location = CONFIG['CHROME_BINARY_PATH']
    options.add_argument(f'--user-data-dir={CONFIG["USER_DATA_DIR"]}')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--log-level=3')
    return options


@contextmanager
def open_browser():
    driver = None
    for attempt in range(1, CONFIG['MAX_RETRIES'] + 1):
        try:
            service = Service(CONFIG['CHROME_DRIVER_PATH'])
            driver = webdriver.Chrome(service=service, options=get_chrome_options())
            logger.info("Chrome WebDriver启动成功。")
            yield driver
            return
        except Exception as e:
            logger.error(f"启动WebDriver失败 (尝试 {attempt}/{CONFIG['MAX_RETRIES']}): {e}")
            if attempt == CONFIG['MAX_RETRIES']:
                raise
            time.sleep(3)
    if driver:
        driver.quit()


def open_alibaba(driver, selected_categories, sheet_names):
    try:
        if driver:
            url = "https://www.alibaba.com/"
            logger.info(f"访问页面: {url}")
            driver.get(url)
            search_bar = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'fy23-icbu-search-bar-inner'))
            )

            total_success_count = 0
            for category in selected_categories:
                try:
                    success_count = process_link(driver, "https://www.alibaba.com/", category, sheet_names)
                    total_success_count += success_count
                except Exception as e:
                    logger.error(f"处理类别 '{category}' 出错: {e}")

            logger.info(f"总共成功导入的产品数量：{total_success_count}")
            driver.quit()

    except NoSuchElementException as e:
        logger.error(f"未找到元素：{e}")
    except TimeoutException as e:
        logger.error(f"超时等待元素加载：{e}")


def process_link(driver, link, category, sheet_name):
    """处理单个链接的主要逻辑"""
    success_count = 0
    try:
        logger.info(f"处理分类: {category}")
        logger.info(f"处理链接: {link}")

        # 导航到链接并等待搜索框加载
        driver.get(link)
        search_input = WebDriverWait(driver, CONFIG['WAIT_TIMEOUT']).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input.search-bar-input.util-ellipsis'))
        )

        # 搜索产品
        search_input.clear()
        search_input.send_keys(category)
        driver.find_element(By.CSS_SELECTOR, 'button.fy23-icbu-search-bar-inner-button').click()

        # 等待产品列表加载
        WebDriverWait(driver, CONFIG['WAIT_TIMEOUT_LONG']).until(
            EC.presence_of_element_located((By.CLASS_NAME, "organic-list"))
        )

        # 滚动加载所有产品
        last_height = driver.execute_script("return document.body.scrollHeight")
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(CONFIG['SCROLL_WAIT'])
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        # 处理产品列表
        product_list = driver.find_elements(By.CLASS_NAME, "fy23-search-card")
        logger.info(f"找到 {len(product_list)} 个产品")

        for product in product_list:
            try:
                # 获取产品标题和链接
                title = product.find_element(By.CLASS_NAME, "search-card-e-title").text
                link = product.find_element(By.TAG_NAME, "a").get_attribute("href")
                logger.info(f"处理产品: {title}")

                # 在新标签页中打开产品
                driver.execute_script(f"window.open('{link}')")
                success_count = handle_product_detail(driver, category, success_count, sheet_name)

            except Exception as e:
                logger.error(f"处理单个产品时出错: {e}")
                continue

        return success_count

    except Exception as e:
        logger.error(f"处理链接时出错: {e}")
        return success_count


def handle_product_detail(driver, category, success_count, sheet_name):
    """处理产品详情页面"""
    original_window = driver.current_window_handle
    new_window = None

    try:
        # 获取新打开的窗口句柄
        new_window = [handle for handle in driver.window_handles
                      if handle != original_window][0]

        # 切换到新窗口
        driver.switch_to.window(new_window)

        try:
            # 等待页面加载
            WebDriverWait(driver, CONFIG['WAIT_TIMEOUT']).until(
                EC.presence_of_element_located((By.TAG_NAME, "h1")))

            # 检查产品是否可发货
            if check_shipping_error(driver):
                logger.info("产品无法发货到当前地区，跳过")
                return success_count

            # 检查产品是否已存在
            if check_product_exists(driver):
                logger.info("产品已存在，跳过")
                return success_count

            # 处理产品导入
            success_count = process_product_import(driver, category, success_count, sheet_name)

        finally:
            # 确保关闭新窗口并切回原窗口
            if new_window in driver.window_handles:
                driver.close()
                driver.switch_to.window(original_window)

        return success_count

    except Exception as e:
        logger.error(f"处理产品详情页时发生错误: {e}")
        # 确保在发生错误时也能正确关闭窗口
        if new_window and new_window in driver.window_handles:
            driver.switch_to.window(new_window)
            driver.close()
            driver.switch_to.window(original_window)
        return success_count


def check_product_exists(driver):
    try:
        message = driver.find_element(
            By.XPATH,
            '//div[@class="textcontainer centeralign home-content "]/p[1]'
        ).text
        return message == "This product is already in your store, what would you like to do?"
    except NoSuchElementException:
        return False


def process_product_import(driver, category, success_count, sheet_name):
    try:
        # 执行导入步骤
        perform_import_steps(driver, sheet_name)

        # 等待导入完成
        if wait_for_import_completion(driver):
            success_count += 1
            logger.info(f"产品导入成功，总数: {success_count}")

        return success_count
    except Exception as e:
        logger.error(f"产品导入过程中发生错误: {e}")
        return success_count


def fetch_dropdown_options(driver, sheet_name):
    """处理下拉菜单选项的选择"""
    try:
        # 确保 sheet_name 是字符串而不是列表
        if isinstance(sheet_name, list):
            sheet_name = sheet_name[0]

        logger.info(f"输入关键词: {sheet_name}")

        # 等待下拉菜单的整个区域可见
        dropdown = WebDriverWait(driver, CONFIG['WAIT_TIMEOUT']).until(
            EC.visibility_of_element_located((By.CLASS_NAME, 'ms-drop')))
        logger.info("找到下拉菜单区域")

        # 找到搜索框并输入关键词
        search_box = dropdown.find_element(By.CSS_SELECTOR, '.ms-search input[type="text"]')
        search_box.clear()
        search_box.send_keys(sheet_name.lower())
        logger.info(f"在搜索框中输入关键词: {sheet_name.lower()}")

        # 等待搜索结果加载完成
        time.sleep(CONFIG['ANIMATION_WAIT'])

        # 使用JavaScript取消所有复选框的选中状态
        driver.execute_script("""
            var checkboxes = document.querySelectorAll('input[data-name="selectItem"]');
            checkboxes.forEach(function(checkbox) {
                if (checkbox.checked) {
                    checkbox.click();
                }
            });
        """)
        logger.info("取消所有复选框的选中状态")

        # 使用JavaScript选中与给定关键词匹配的复选框
        driver.execute_script("""
            var checkboxes = document.querySelectorAll('input[data-name="selectItem"]');
            var searchTerm = arguments[0].toLowerCase();
            checkboxes.forEach(function(checkbox) {
                var spanElement = checkbox.nextElementSibling;
                if (spanElement && spanElement.innerText.toLowerCase() === searchTerm) {
                    checkbox.click();
                }
            });
        """, sheet_name.lower())
        logger.info(f"选中匹配关键词的复选框: {sheet_name}")

    except TimeoutException:
        logger.error("超时：无法加载下拉菜单或搜索结果")
    except Exception as e:
        logger.error(f"处理下拉选项时出错: {e}")


def handle_product_actions(browser, category, success_count, sheet_name):
    logger.info(f"处理产品详情页操作: {category}, {sheet_name}!!!")
    try:
        add_btn_con = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="addBtnCon"]')))
        add_btn_con.click()
        logging.info("点击了按钮//*[@id='addBtnCon']")

        try:
            element = WebDriverWait(browser, 20).until(
                EC.presence_of_element_located((By.XPATH, '//span[@class="inactive" and text()="Draft"]'))
            )
            logging.info("成功加载 Draft 元素")
            actions = ActionChains(browser)
            actions.move_to_element(element).perform()
            element.click()
            logging.info("成功点击 Draft 元素")
            time.sleep(2)
        except Exception as e:
            logging.error(f"等待和点击 Draft 元素时出现错误：{e}")
            close_current_tab(browser)
            return success_count

        time.sleep(3)  # 可以根据实际情况调整等待时间

        # 等待 "Sorry, this product can't be shipped to your region." 元素出现
        try:
            WebDriverWait(browser, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//div[contains(text(), "Sorry, this product can\'t be shipped to your region.")]'))
            )
            logging.info("'检测到产品无法配送到当前区域，跳过'处理。")
            browser.close()  # 关闭当前产品详情页标签页
            return success_count  # 返回 success_count，继续处理下一款产品
        except TimeoutException:
            logging.info("未检测到区域限制消息，继续处理。")
            pass  # 如果未找到消息元素，继续后续操作

        # 检查是否出现 "This product is already in your store, what would you like to do?"
        success_message = None
        try:
            success_message = browser.find_element(By.XPATH,
                                                   '//div[@class="textcontainer centeralign home-content "]/p[1]')
            if success_message.text == "This product is already in your store, what would you like to do?":
                logging.info("产品已存在，不再处理当前产品")
                browser.close()  # 关闭当前产品详情页标签页
                return success_count  # 跳出函数，不再处理当前产品
        except NoSuchElementException:
            pass  # 如果未找到消息元素，继续后续操作

        time.sleep(2)  # 可以根据实际情况调整等待时间

        # 继续后续操作，例如选择下拉菜单中的类别等
        select_button = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@class="ms-choice"]'))
        )
        logging.info("等待并点击选择按钮")
        select_button.click()

        dropdown = WebDriverWait(browser, 10).until(
            EC.visibility_of_element_located((By.CLASS_NAME, "ms-drop"))
        )

        fetch_dropdown_options(browser, sheet_name)
        time.sleep(3)

        try:
            # 点击 description_tab_button 按钮
            description_tab_button = browser.find_element(By.XPATH, '//*[@id="description_tab_button"]')
            description_tab_button.click()
            logging.info("点击了 description_tab_button 按钮")
            time.sleep(3)  # 等待页面加载

            # 点击 Variants 按钮
            variants_button = browser.find_element(By.CSS_SELECTOR,
                                                   'button.accordion-tab[data-actab-group="0"][data-actab-id="2"]')
            variants_button.click()
            logging.info("点击了 Variants 按钮")

            # 选择 Import all variants automatically 单选框
            all_variants_radio = browser.find_element(By.ID, 'all_variants')
            all_variants_radio.click()
            logging.info("选择 Import all variants automatically 单选框")

            time.sleep(3)  # 等待页面反应

            # 选择 Select which variants to include 单选框
            price_switch_radio = browser.find_element(By.ID, 'price_switch')
            price_switch_radio.click()
            logging.info("选择 Select which variants to include 单选框")

            time.sleep(3)  # 等待页面反应
        except Exception as e:
            logging.error(f"点击 Variants 按钮时出现错误：{e}")
            close_current_tab(browser)
            return success_count

        # 点击 Images 按钮
        images_button = browser.find_element(By.XPATH,
                                             '//button[@class="accordion-tab accordion-custom-tab" and @data-actab-group="0" and @data-actab-id="3"]')
        images_button.click()
        logging.info("点击了 Images 按钮")
        time.sleep(3)  # 等待页面反应

        add_to_store_button = browser.find_element(By.ID, 'addBtnSec')
        scroll_to_element(browser, add_to_store_button)

        add_to_store_button.click()
        logging.info("成功点击 Add to your Store 按钮")

        logging.info("等待页面加载完成")

        # 等待导入过程完成，确保 importify-app-container 元素出现
        try:
            wait_for_element_to_appear(browser, By.ID, 'importify-app-container')
            logging.info("产品正在导入中...")

            # 等待成功消息出现
            success_message = None
            timeout = 100  # 设定超时时间
            start_time = time.time()
            while time.time() - start_time < timeout:
                try:
                    success_message = browser.find_element(By.XPATH,
                                                           '//div[@class="textcontainer centeralign home-content "]/p[1]')
                    if success_message.text == "We have successfully created the product page.":
                        logging.info(f"产品导入成功, 共计: {success_count + 1}")
                        success_count += 1
                        break
                    else:
                        logging.warning("产品正在导入中...")
                except Exception as e:
                    logging.warning("未检测到产品成功导入，继续等待...")
                time.sleep(5)  # 每秒检查一次

            if not success_message or success_message.text != "We have successfully created the product page.":
                logging.error("超时：未找到成功创建产品页面的消息")

        except Exception as e:
            logging.error(f"页面加载出错: {e}")
            close_current_tab(browser)

        time.sleep(3)
        browser.close()
        return success_count

    except NoSuchWindowException as e:
        logging.error(f"浏览器窗口丢失：{e}")
        return success_count
    except Exception as e:
        logging.error(f"处理产品详情页操作时发生错误: {e}")
        return success_count


def check_shipping_error(driver):
    """
    检查产品详情页中是否有与无法发货相关的错误消息
    """
    try:
        # 使用新的 XPath 查找包含错误消息的元素
        error_message = driver.find_element(By.XPATH, '//div[@class="unsafe-unableToShip"]')

        # 检查元素是否显示在页面上
        if error_message.is_displayed():
            return True
    except NoSuchElementException:
        # 如果未找到元素，则返回 False
        return False

    # 默认情况下返回 False
    return False


def close_current_tab(browser):
    try:
        if len(browser.window_handles) > 1:
            # 关闭当前标签页
            browser.close()
            # 切换到最后一个标签页
            browser.switch_to.window(browser.window_handles[-1])
        else:
            # 如果只有一个标签页，则关闭它
            browser.close()
            logging.info("所有标签页已关闭，准备处理下一个产品")
    except NoSuchWindowException as e:
        logging.error(f"浏览器窗口丢失：{e}")
    except Exception as e:
        logging.error(f"关闭标签页时发生错误: {e}")


def wait_for_element_to_appear(driver, by, selector, timeout=10):
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, selector))
        )
    except TimeoutException:
        logging.error(f"元素未能在 {timeout} 秒内出现: {selector}")
        raise


def close_tab(driver, window_handle):
    """关闭指定的浏览器标签页"""
    try:
        if window_handle:
            driver.switch_to.window(window_handle)
            driver.close()
            logger.info("关闭标签页")
            # 切换回主窗口
            if len(driver.window_handles) > 0:
                driver.switch_to.window(driver.window_handles[0])
    except Exception as e:
        logger.error(f"关闭标签页时发生错误: {e}")


def get_screen_width():
    try:
        root = tk.Tk()
        screen_width = root.winfo_screenwidth()
        root.destroy()
        return screen_width
    except Exception as e:
        logger.error(f"获取屏幕宽度时出错: {e}")
        return 1000  # 返回默认屏幕宽度


def browse_excel_file():
    root = tk.Tk()
    root.withdraw()  # 隐藏Tk窗口
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    return file_path


def read_categories_from_excel(file_path):
    try:
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        categories = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            category = row[0]
            if category:
                categories.append(category)
        return categories
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return []


def read_sheet_names_from_excel(file_path):
    sheet_name = []
    try:
        wb = load_workbook(filename=file_path)
        sheet_name = wb.sheetnames
    except Exception as e:
        logger.error(f"读取Excel文件时发生错误: {e}")
    return sheet_name


def scroll_to_element(driver, element):
    """优化的元素滚动函数"""
    try:
        # 使用显式等待确保元素可见
        WebDriverWait(driver, CONFIG['WAIT_TIMEOUT']).until(EC.visibility_of(element))

        # 使用平滑滚动
        driver.execute_script(
            "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
            element
        )

        # 等待滚动动画完成
        time.sleep(CONFIG['ANIMATION_WAIT'])

        logger.info(f"滚动到元素: {element.text if hasattr(element, 'text') else '未知元素'}")
    except Exception as e:
        logger.error(f"滚动到元素时出错: {e}")


def perform_import_steps(driver, sheet_name):
    """执行产品导入的具体步骤，优化等待时间"""
    try:
        # 使用显式等待替代固定时间等待
        wait = WebDriverWait(driver, CONFIG['WAIT_TIMEOUT'])
        wait_long = WebDriverWait(driver, CONFIG['WAIT_TIMEOUT_LONG'])

        # 点击添加按钮
        add_btn = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="addBtnCon"]')))
        add_btn.click()
        logger.info("点击了添加按钮")

        # 点击 Draft - 使用长等待时间
        draft_element = wait_long.until(
            EC.element_to_be_clickable((By.XPATH, '//span[@class="inactive" and text()="Draft"]')))
        ActionChains(driver).move_to_element(draft_element).click().perform()
        logger.info("点击了 Draft 选项")

        # 等待页面加载完成
        wait.until(
            EC.presence_of_element_located((By.XPATH, '//button[@class="ms-choice"]')))

        # 选择类别
        select_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//button[@class="ms-choice"]')))
        select_button.click()
        logger.info("打开类别选择下拉框")

        # 处理下拉选项
        fetch_dropdown_options(driver, sheet_name)
        time.sleep(CONFIG['ANIMATION_WAIT'])  # 仅等待动画完成

        # 使用自定义等待条件检查元素可交互
        def element_is_ready(driver, xpath):
            element = driver.find_element(By.XPATH, xpath)
            return element.is_displayed() and element.is_enabled()

        # 处理描述标签
        wait.until(lambda d: element_is_ready(d, '//*[@id="description_tab_button"]'))
        description_tab = driver.find_element(By.XPATH, '//*[@id="description_tab_button"]')
        description_tab.click()
        logger.info("点击了描述标签")

        # 处理变体 - 使用显式等待替代固定等待
        variants_button = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR,
                                        'button.accordion-tab[data-actab-group="0"][data-actab-id="2"]')))
        variants_button.click()
        logger.info("点击了变体按钮")

        # 选择变体选项
        wait.until(EC.element_to_be_clickable((By.ID, 'all_variants'))).click()
        time.sleep(CONFIG['ANIMATION_WAIT'])  # 等待动画完成

        wait.until(EC.element_to_be_clickable((By.ID, 'price_switch'))).click()
        time.sleep(CONFIG['ANIMATION_WAIT'])  # 等待动画完成

        # 处理图片
        images_button = wait.until(
            EC.element_to_be_clickable((By.XPATH,
                                        '//button[@class="accordion-tab accordion-custom-tab" and @data-actab-group="0" and @data-actab-id="3"]')))
        images_button.click()
        logger.info("点击了图片按钮")

        # 点击添加到商店
        add_to_store = wait.until(EC.element_to_be_clickable((By.ID, 'addBtnSec')))
        scroll_to_element(driver, add_to_store)
        add_to_store.click()
        logger.info("点击了添加到商店按钮")

    except Exception as e:
        logger.error(f"执行导入步骤时出错: {e}")
        raise


def wait_for_import_completion(driver, timeout=None):
    """优化的导入完成等待函数"""
    if timeout is None:
        timeout = CONFIG['IMPORT_TIMEOUT']

    try:
        # 等待导入容器出现
        WebDriverWait(driver, CONFIG['WAIT_TIMEOUT']).until(
            EC.presence_of_element_located((By.ID, 'importify-app-container')))
        logger.info("产品导入进行中...")

        # 使用显式等待检查成功消息
        success_xpath = '//div[@class="textcontainer centeralign home-content "]/p[1]'
        success_condition = EC.text_to_be_present_in_element(
            (By.XPATH, success_xpath),
            "We have successfully created the product page."
        )

        try:
            WebDriverWait(driver, timeout).until(success_condition)
            logger.info("产品导入成功")
            return True
        except TimeoutException:
            logger.warning(f"导入超时（{timeout}秒）")
            return False

    except Exception as e:
        logger.error(f"等待导入完成时出错: {e}")
        return False


def main():
    try:
        file_path = browse_excel_file()
        if not file_path:
            logger.error("未选择Excel文件。")
            return

        selected_categories = read_categories_from_excel(file_path)
        if not selected_categories:
            logger.error("未从Excel文件中读取到任何类别。")
            return
        logger.info(f"从Excel文件中读取的要导入的产品名称: {selected_categories}")

        with open_browser() as driver:
            if not driver:
                logger.error("无法启动浏览器。")
                return

            # 获取工作表名称列表
            sheet_name = read_sheet_names_from_excel(file_path)
            if not sheet_name:
                logger.error("未从Excel文件中读取到任何工作表名称。")
                return
            logger.info(f"从Excel文件中读取的工作表名称: {sheet_name}")

            # 调用 open_alibaba() 函数，并传递 driver、selected_categories 和 sheet_names
            open_alibaba(driver, selected_categories, sheet_name)

    except Exception as e:
        pass
    input("已完成所有内容")


if __name__ == "__main__":
    main()
